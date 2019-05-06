# https://openpyxl.readthedocs.io
from openpyxl import Workbook
from openpyxl import load_workbook
from collections import namedtuple


def get_data(filename):
    wb = load_workbook(filename=filename)
    raw_data = []

    for row in wb['Form1']:
        r = [str(x.value) for x in row]
        raw_data.append(r)

    headers = list(map(lambda x: x.replace(" ", ""), raw_data[0]))
    row = namedtuple('row', headers)

    return [row(*x) for x in raw_data[1:]]


def write_data(data):
    filename = 'output.xlsx'
    try:
        wb = load_workbook(filename=filename)

    except IOError:
        wb = Workbook()

    ws = wb.active
    ws.title = 'jobs'

    headers = ['date', 'name', 'email', 'school', 'create_blueprint',
               'use_blueprint', 'associations', 'checksum', 'computed_blueprint', 'computed_associations', 'message']
    for i, j in enumerate(headers):
        ws.cell(row=1, column=i+1, value=j)

    row_count = 1
    for row in data:
        print(row)
        this_row = row['who']
        blueprint = row['blueprint']['course_id']
        associations = ",".join(list(map(lambda x: x['course_id'], row['associations'])))
        this_row.update(dict(checksum=row['checksum'], computed_blueprint=blueprint, computed_associations=associations,  message=",".join(row['error'])))
        row_count += 1
        for i, j in enumerate(this_row.values()):
            print((row_count, i+1, j))
            ws.cell(row=row_count, column=i+1, value=j)

    wb.save(filename=filename)


if __name__ == '__main__':
    # data_file_name = 'Blueprint Course Management 1 (14).xlsx'
    # print(get_data(data_file_name))

    write_data('x')
